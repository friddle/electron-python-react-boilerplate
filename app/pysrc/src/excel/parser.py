# encoding:utf-8

import os
from openpyxl import load_workbook

from data.ClassDetail import ClassDetail
import threading
from functools import partial


class Parser:
    def __init__(self):
        self.import_task = None

    def import_excel(self, path):
        if not os.path.exists(path):
            return False
        wb = load_workbook(filename=path, read_only=False, guess_types=True)
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            self.__import_class_detail_excel__(sheet, sheet_name)
        return True

    def import_excel_async(self, path):
        self.import_task = threading.Thread(target=self.import_excel, args=[str(path)])
        self.import_task.start()
        return True

    def status_import(self):
        if self.import_task.is_alive() is False:
            return "finish"
        else:
            return "running"

    def __import_class_detail_excel__(self, sheet, sheet_name):
        for row in range(2, sheet.max_row):
            if sheet_name != u"班课" and sheet_name != u"一对一":
                return
            detail_column = self.__covert_class_data__(sheet, row)
            if sheet_name == u"一对一":
                detail_column.class_type = "一对一"
            if detail_column.additions is None:
                detail_column.additions = ""
            try:
                detail_column.save()
            except:
                print(row)

    def __covert_class_data__(self, sheet, row):
        detail = ClassDetail()
        detail.date = sheet[row][0].value
        detail.range = sheet[row][2].value
        detail.student = sheet[row][3].value
        detail.gradle = sheet[row][4].value
        detail.subject = sheet[row][5].value
        detail.teacher = sheet[row][6].value
        detail.teacher_type = sheet[row][7].value
        detail.manager = sheet[row][8].value
        detail.class_times = sheet[row][9].value
        detail.class_type = sheet[row][10].value
        detail.additions = sheet[row][11].value
        detail.finished = True
        return detail

    def output_excel(self, filename, database_obj):
        print("output excel")
        return "output"


if __name__ == "__main__":
    parser = Parser()
    parser.import_excel("./import.xlsx", "ClassDetail")
